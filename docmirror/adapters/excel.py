"""
Excel Adapter — Excel → BaseResult

使用 openpyxl 提取工作表数据为表格 Block。
"""

from __future__ import annotations

import logging
from pathlib import Path

from docmirror.framework.base import BaseParser, ParserOutput, ParserStatus
from docmirror.models.domain import BaseResult, Block, PageLayout

logger = logging.getLogger(__name__)


class ExcelAdapter(BaseParser):
    """Excel (.xlsx) 格式适配器。"""

    async def to_base_result(self, file_path: Path) -> BaseResult:
        """Excel → BaseResult (每个 sheet 对应一个 PageLayout)。"""
        import openpyxl
        wb = openpyxl.load_workbook(str(file_path), data_only=True)

        pages = []
        text_parts = []

        for idx, sheet_name in enumerate(wb.sheetnames):
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row_data = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in row_data):
                    rows.append(row_data)

            if not rows:
                continue

            block = Block(
                block_type="table",
                raw_content=rows,
                page=idx,
            )
            title_block = Block(
                block_type="title",
                raw_content=f"Sheet: {sheet_name}",
                page=idx,
                heading_level=2,
            )
            pages.append(PageLayout(
                page_number=idx,
                blocks=(title_block, block),
            ))
            text_parts.append(f"## {sheet_name}\n" + "\n".join(" | ".join(r) for r in rows[:5]))

        return BaseResult(
            pages=tuple(pages),
            full_text="\n\n".join(text_parts),
            metadata={"source_format": "excel", "sheet_names": wb.sheetnames},
        )


