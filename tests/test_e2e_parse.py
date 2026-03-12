"""End-to-end batch parsing test for all files under tests/fixtures.

Recursively discovers every file in the fixtures directory, parses each
through ParserDispatcher, and writes a consolidated Excel report.
"""
import asyncio
import json
import logging
import os
import sys
import time
from pathlib import Path

os.environ.setdefault("NO_PROXY", "localhost,127.0.0.1")

# Enable INFO logging for docmirror (shows per-page timing breakdown)
logging.basicConfig(
    level=logging.WARNING,
    format="%(message)s",
)
logging.getLogger("docmirror.core.extraction.extractor").setLevel(logging.INFO)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from docmirror.framework.dispatcher import ParserDispatcher
from docmirror.models.document_types import DocumentType


# ────────────────────────────────────────────────────────────
# Pretty CLI progress display
# ────────────────────────────────────────────────────────────

STEP_ICONS = {
    "Validating file":    "📁",
    "Checking cache":     "💾",
    "Detecting file type": "🔍",
    "Security scan":      "🔒",
    "Extracting content": "⚙️ ",
}

_step_times: dict[int, float] = {}
_current_file: str = ""


def cli_progress(step: int, total: int, name: str, detail: str):
    """Pretty-print each pipeline step as it happens."""
    icon = STEP_ICONS.get(name, "▸")

    now = time.time()
    if step not in _step_times:
        _step_times[step] = now
    elapsed = now - _step_times[step]

    bar = f"[{step}/{total}]"
    time_str = f"({elapsed:.1f}s)" if elapsed > 0.05 else ""
    print(f"  {bar} {icon} {name:<22} {detail} {time_str}", flush=True)


# ────────────────────────────────────────────────────────────
# File discovery
# ────────────────────────────────────────────────────────────

SKIP_NAMES = {".DS_Store", ".gitkeep", "Thumbs.db"}


def discover_files(root: Path) -> list[Path]:
    """Recursively collect all parsable files under *root*."""
    files: list[Path] = []
    for path in sorted(root.rglob("*")):
        if path.is_file() and path.name not in SKIP_NAMES:
            files.append(path)
    return files


# ────────────────────────────────────────────────────────────
# Excel report generation
# ────────────────────────────────────────────────────────────

# Style constants
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUCCESS_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def _style_header_row(ws, col_count: int):
    """Apply uniform header styling."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws, min_width: int = 10, max_width: int = 60):
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = min_width
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


def write_excel_report(
    results: list[dict],
    output_path: Path,
    fixtures_root: Path,
):
    """Write the batch-parse results into a formatted Excel workbook.

    Sheets created:
      1. **Summary** – one row per file with status, confidence, text length, etc.
      2. **Tables**  – flattened view of every extracted table across all files.
      3. **Entities** – flattened view of extracted entities.
      4. **Errors**  – files that failed parsing.
    """
    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    summary_headers = [
        "No.", "Folder", "File Name", "Status", "Confidence",
        "Text Length", "Table Count", "Elapsed (ms)",
        "Text Preview (200 chars)", "Error",
    ]
    ws_summary.append(summary_headers)
    _style_header_row(ws_summary, len(summary_headers))

    for idx, r in enumerate(results, 1):
        rel = r["relative_path"]
        folder = str(Path(rel).parent) if str(Path(rel).parent) != "." else ""
        row = [
            idx,
            folder,
            r["file_name"],
            r["status"],
            r.get("confidence"),
            r.get("text_length"),
            r.get("table_count", 0),
            r.get("elapsed_ms"),
            (r.get("text_preview") or "")[:200],
            r.get("error", ""),
        ]
        ws_summary.append(row)
        row_num = idx + 1
        fill = SUCCESS_FILL if r.get("success") else ERROR_FILL
        for col in range(1, len(summary_headers) + 1):
            cell = ws_summary.cell(row=row_num, column=col)
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            if col in (4,):  # Status column
                cell.fill = fill

    ws_summary.auto_filter.ref = ws_summary.dimensions
    ws_summary.freeze_panes = "A2"
    _auto_width(ws_summary)

    # ── Sheet 2: Tables ───────────────────────────────────
    ws_tables = wb.create_sheet("Tables")
    table_headers = ["File Name", "Table #", "Rows", "Cols", "Header Row", "Sample Row 1"]
    ws_tables.append(table_headers)
    _style_header_row(ws_tables, len(table_headers))

    table_row_idx = 1
    for r in results:
        for td in r.get("tables_detail", []):
            table_row_idx += 1
            header_str = " | ".join(str(c) for c in td.get("header", []))
            sample_str = " | ".join(str(c) for c in td.get("row_1", []))
            ws_tables.append([
                r["file_name"],
                td.get("table_index", ""),
                td.get("rows"),
                td.get("cols"),
                header_str,
                sample_str,
            ])
            for col in range(1, len(table_headers) + 1):
                cell = ws_tables.cell(row=table_row_idx + 1, column=col)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN

    ws_tables.auto_filter.ref = ws_tables.dimensions
    ws_tables.freeze_panes = "A2"
    _auto_width(ws_tables)

    # ── Sheet 3: Entities ─────────────────────────────────
    ws_entities = wb.create_sheet("Entities")
    entity_headers = ["File Name", "Entity Key", "Entity Value"]
    ws_entities.append(entity_headers)
    _style_header_row(ws_entities, len(entity_headers))

    ent_row = 1
    for r in results:
        for k, v in r.get("entities", {}).items():
            ent_row += 1
            ws_entities.append([r["file_name"], k, str(v)])
            for col in range(1, len(entity_headers) + 1):
                cell = ws_entities.cell(row=ent_row + 1, column=col)
                cell.border = THIN_BORDER
                cell.alignment = WRAP_ALIGN

    ws_entities.auto_filter.ref = ws_entities.dimensions
    ws_entities.freeze_panes = "A2"
    _auto_width(ws_entities)

    # ── Sheet 4: Errors ───────────────────────────────────
    ws_errors = wb.create_sheet("Errors")
    error_headers = ["No.", "File Name", "Folder", "Error Message"]
    ws_errors.append(error_headers)
    _style_header_row(ws_errors, len(error_headers))

    err_idx = 0
    for r in results:
        if r.get("error"):
            err_idx += 1
            rel = r["relative_path"]
            folder = str(Path(rel).parent) if str(Path(rel).parent) != "." else ""
            ws_errors.append([err_idx, r["file_name"], folder, r["error"]])
            for col in range(1, len(error_headers) + 1):
                cell = ws_errors.cell(row=err_idx + 1, column=col)
                cell.border = THIN_BORDER
                cell.fill = ERROR_FILL
                cell.alignment = WRAP_ALIGN

    ws_errors.auto_filter.ref = ws_errors.dimensions
    ws_errors.freeze_panes = "A2"
    _auto_width(ws_errors)

    wb.save(output_path)


# ────────────────────────────────────────────────────────────
# Result serialization helper
# ────────────────────────────────────────────────────────────

def _collect_result(result, file_path: Path, fixtures_root: Path) -> dict:
    """Convert a PerceptionResult into a flat dict for the Excel report."""
    rel_path = file_path.relative_to(fixtures_root)

    out: dict = {
        "file_name": file_path.name,
        "relative_path": str(rel_path),
        "status": str(result.status),
        "success": result.success,
        "confidence": result.confidence,
        "text_length": len(result.content.text),
        "text_preview": result.content.text[:500],
        "table_count": len(result.tables),
        "elapsed_ms": result.timing.elapsed_ms if result.timing else None,
        "error": result.error.message if result.error else None,
        "entities": {},
        "tables_detail": [],
    }

    if result.content.entities:
        out["entities"] = dict(result.content.entities)

    for i, t in enumerate(result.tables):
        if hasattr(t, "data") and t.data:
            table_data = t.data
        elif isinstance(t, list) and t:
            table_data = t
        else:
            continue
        rows = len(table_data)
        cols = len(table_data[0]) if table_data else 0
        td: dict = {"table_index": i, "rows": rows, "cols": cols}
        if table_data:
            td["header"] = table_data[0]
            if rows > 1:
                td["row_1"] = table_data[1]
        out["tables_detail"].append(td)

    return out


# ────────────────────────────────────────────────────────────
# Main entry point
# ────────────────────────────────────────────────────────────

async def main():
    import argparse

    parser = argparse.ArgumentParser(
        description=(
            "Batch end-to-end parsing of all files under tests/fixtures. "
            "Results are written to an Excel report."
        ),
    )
    parser.add_argument(
        "--fixtures-dir", default=None,
        help="Root fixtures directory (default: tests/fixtures)",
    )
    parser.add_argument(
        "--output", "-o", default=None,
        help="Output Excel path (default: tests/e2e_parse_results.xlsx)",
    )
    parser.add_argument(
        "--skip-cache", action="store_true",
        help="Skip cache, force re-parsing",
    )
    parser.add_argument(
        "--file", default=None,
        help="Parse a single file instead of batching all fixtures",
    )
    args = parser.parse_args()

    fixtures_root = Path(args.fixtures_dir).resolve() if args.fixtures_dir else (
        Path(__file__).parent / "fixtures"
    )
    output_path = Path(args.output) if args.output else (
        Path(__file__).parent / "e2e_parse_results.xlsx"
    )

    # ── Single-file mode (backward compatible) ──
    if args.file:
        files = [Path(args.file).resolve()]
        if not files[0].exists():
            print(f"ERROR: {files[0]} not found")
            return
    else:
        files = discover_files(fixtures_root)
        if not files:
            print(f"ERROR: No files found under {fixtures_root}")
            return

    total_files = len(files)
    print(f"\n{'=' * 60}")
    print(f"  📂 Batch E2E Parse: {total_files} file(s)")
    print(f"     Fixtures root : {fixtures_root}")
    print(f"     Output Excel  : {output_path}")
    print(f"{'=' * 60}\n")

    dispatcher = ParserDispatcher()
    all_results: list[dict] = []
    success_count = 0
    fail_count = 0

    for idx, file_path in enumerate(files, 1):
        rel = file_path.relative_to(fixtures_root) if file_path.is_relative_to(fixtures_root) else file_path.name
        size = file_path.stat().st_size

        print(f"\n┌─ [{idx}/{total_files}] {rel} ({size:,} bytes)")
        print(f"│")

        _step_times.clear()
        t0 = time.time()

        try:
            result = await dispatcher.process(
                str(file_path),
                document_type=DocumentType.OTHER,
                skip_cache=args.skip_cache,
                on_progress=cli_progress,
            )

            elapsed = (time.time() - t0) * 1000
            status_icon = "✅" if result.success else "❌"
            print(f"│")
            print(f"│  {status_icon} Status: {result.status}  "
                  f"Confidence: {result.confidence:.4f}  "
                  f"Text: {len(result.content.text):,} chars  "
                  f"Tables: {len(result.tables)}  "
                  f"Time: {elapsed:.0f}ms")

            if result.error:
                print(f"│  ⚠️  Error: {result.error.message}")
                fail_count += 1
            else:
                success_count += 1

            collected = _collect_result(result, file_path, fixtures_root)
            all_results.append(collected)

        except Exception as e:
            fail_count += 1
            print(f"│  💥 Exception: {e}")
            all_results.append({
                "file_name": file_path.name,
                "relative_path": str(rel),
                "status": "EXCEPTION",
                "success": False,
                "confidence": 0.0,
                "text_length": 0,
                "text_preview": "",
                "table_count": 0,
                "elapsed_ms": (time.time() - t0) * 1000,
                "error": str(e),
                "entities": {},
                "tables_detail": [],
            })

        print(f"└{'─' * 59}")

    # ── Write Excel report ──
    print(f"\n{'=' * 60}")
    print(f"  📊 Writing Excel report...")
    write_excel_report(all_results, output_path, fixtures_root)
    print(f"  💾 Saved to: {output_path}")

    # ── Final summary ──
    print(f"\n  ── Summary ──")
    print(f"     Total files : {total_files}")
    print(f"     ✅ Success  : {success_count}")
    print(f"     ❌ Failed   : {fail_count}")
    print(f"{'=' * 60}\n")

    # Also save a JSON sidecar for programmatic access
    json_path = output_path.with_suffix(".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2, default=str)
    print(f"  📄 JSON sidecar: {json_path}")


if __name__ == "__main__":
    asyncio.run(main())
